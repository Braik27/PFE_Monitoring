"""
engine/country_detail_report.py — Rapport détaillé LIGNE par pays.

Principe :
  Le rapport "Excel analyse" (engine/detailed_report.py) produit une ligne
  par article. Chaque ligne du DataFrame `merged` (outer merge Cegid/Oracle)
  peut porter un OPERATING_UNIT_CODE — colonne présente dans les fichiers
  réels mais absente des fixtures et du registry. On lit ce code sur chaque
  ligne (côté Cegid d'abord, sinon Oracle), on le traduit en pays via la
  table validée par le client, et on écrit UN fichier Excel détaillé par
  pays — même structure exacte que l'Excel analyse (Résumé / Rapport
  détaillé / Écarts), via export_detailed_excel sur le sous-ensemble.

Mapping OPERATING_UNIT_CODE → pays (source de vérité : documentation ABA,
validée utilisateur — ne pas modifier sans validation) :

    Qatar  : SPG, LUX, LNH, FRT, ISM, DID
    Koweït : RMK, KLD
    KSA    : PSC

Tout autre code, ou ligne sans code → bucket AUTRE ("Autre / Non classé"),
jamais perdu silencieusement.
"""
from __future__ import annotations
import logging
from typing import Optional

from pandas import isna as _pd_isna

log = logging.getLogger(__name__)

# ── Mapping client ABA ───────────────────────────────────────────────────────
OU_COUNTRY_MAP: dict[str, str] = {
    # Qatar
    "SPG": "QATAR", "LUX": "QATAR", "LNH": "QATAR",
    "FRT": "QATAR", "ISM": "QATAR", "DID": "QATAR",
    # Koweït
    "RMK": "KUWAIT", "KLD": "KUWAIT",
    # Arabie Saoudite
    "PSC": "KSA",
}

COUNTRY_LABELS: dict[str, str] = {
    "QATAR": "🇶🇦 Qatar", "KUWAIT": "🇰🇼 Kuwait", "KSA": "🇸🇦 KSA",
}
AUTRE = "AUTRE"
AUTRE_LABEL = "❔ Autre / Non classé"
_FILENAME_BY_COUNTRY: dict[str, str] = {
    "QATAR": "qatar", "KUWAIT": "kuwait", "KSA": "ksa", AUTRE: "autre",
}

# Anciens buckets stockés dans les résumés (ancien detector) → pays
LEGACY_BUCKET_COUNTRY: dict[str, str] = {
    "LUX": "QATAR", "DOHA": "QATAR", "DAW7A": "QATAR", "SPG": "QATAR",
    "KWT": "KUWAIT",
    "KSA": "KSA",
}


def country_of_ou(code) -> str:
    """Traduit un code OU brut en pays ; inconnu/vide → AUTRE."""
    if code is None:
        return AUTRE
    c = str(code).upper().strip()
    return OU_COUNTRY_MAP.get(c, AUTRE)


def ou_of_merged_row(row) -> Optional[str]:
    """
    Extrait l'OPERATING_UNIT_CODE d'une ligne du merged.
    Priorité côté Cegid (`_cegid`) puis Oracle (`_oracle`).
    Retourne None si absent des deux côtés.
    """
    for col in ("OPERATING_UNIT_CODE_cegid", "OPERATING_UNIT_CODE_oracle"):
        val = row.get(col)
        if val is None:
            continue
        try:
            if _pd_isna(val):       # cellule vide → NaN dans un DataFrame
                continue
        except (TypeError, ValueError):
            pass                    # valeurs non-scalaires : traiter telles quelles
        s = str(val).strip()
        if s:
            return s
    return None


def subset_stats(rows: list[dict]) -> dict[str, int]:
    """
    Compteurs affichés dans l'onglet Résumé, dérivés du sous-ensemble :
    présence d'un numéro de ligne = la ligne venait de ce fichier.
    """
    return {
        "nb_lignes_cegid":  sum(1 for r in rows if r.get("LIGNE_CEGID") is not None),
        "nb_lignes_oracle": sum(1 for r in rows if r.get("LIGNE_ORACLE") is not None),
    }


def read_country_rows(xlsx_path: str) -> list[dict]:
    """
    Relit les lignes de l'onglet "Rapport détaillé" d'un fichier pays
    stocké (même structure que l'Excel analyse).
    Retourne une liste de dicts indexés par en-tête ; [] si absent.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        log.warning("[COUNTRY] Lecture impossible de %s : %s", xlsx_path, e)
        return []

    try:
        if "Rapport détaillé" not in wb.sheetnames:
            log.warning("[COUNTRY] %s : onglet 'Rapport détaillé' absent", xlsx_path)
            return []
        ws = wb["Rapport détaillé"]
        it = ws.iter_rows(values_only=True)
        headers = next(it, None)
        if not headers:
            return []
        out: list[dict] = []
        for raw in it:
            if not any(v is not None and str(v) != "" for v in raw):
                continue                      # ligne vide éventuelle
            out.append(dict(zip(headers, raw)))
        return out
    finally:
        wb.close()


def collect_day_rows(
    analysis_file_maps: list[tuple],
) -> tuple[dict[str, list[dict]], list]:
    """
    Fusionne les lignes de TOUTES les analyses d'une journée, par pays.

    Args:
        analysis_file_maps: liste de (analysis_id, {pays: chemin_xlsx})
                            — chemins bruts depuis summary.country_excel_paths.

    Returns:
        ({pays: [lignes de toutes les sources]}, [ids ignorés])
        Une analyse est ignorée (loggée par l'appelant) si elle n'a pas de
        fichiers stockés exploitables — jamais crash ni perte silencieuse.
    """
    import os

    groups: dict[str, list[dict]] = {}
    skipped: list = []

    for aid, paths_map in analysis_file_maps:
        avail = {c: p for c, p in (paths_map or {}).items()
                 if p and os.path.exists(p)}
        if not avail:
            skipped.append(aid)
            continue
        for country, path in avail.items():
            groups.setdefault(country, []).extend(read_country_rows(path))

    return groups, skipped


def build_country_reports(
    merged,
    report: list[dict],
    *,
    flux_id: str,
    stats: dict | None = None,
    comparison_rules: list[dict] | None = None,
    output_dir: str,
    date_str: str,
) -> dict[str, str]:
    """
    Découpe `report` (sortie de build_detail_report, MÊME ordre que
    merged.iterrows()) par pays et écrit un Excel détaillé par pays.

    Args:
        merged:           DataFrame outer merge (colonnes *_cegid/*_oracle).
        report:           lignes détaillées (build_detail_report(merged,...)).
        flux_id:          identifiant du flux.
        stats:            stats globales (nb_lignes_* pour l'onglet Résumé).
        comparison_rules: règles du registry (sévérités).
        output_dir:       dossier de sortie.
        date_str:         date AAAA-MM-JJ pour les noms de fichiers.

    Returns:
        {pays: chemin_fichier} — uniquement les pays ayant au moins 1 ligne.
    """
    import os

    if len(report) != len(merged):
        log.warning("[COUNTRY] report(%d) et merged(%d) désynchronisés — abandon",
                    len(report), len(merged))
        return {}

    from engine.detailed_report import export_detailed_excel

    # Pays de chaque ligne — merged.iterrows() parcourt dans le même ordre
    # que build_detail_report a construit `report`.
    countries: list[str] = []
    for _, row in merged.iterrows():
        countries.append(country_of_ou(ou_of_merged_row(row)))

    # Regroupement (une ligne = un seul pays)
    groups: dict[str, list[dict]] = {}
    for r, c in zip(report, countries):
        groups.setdefault(c, []).append(r)

    if AUTRE in groups:
        log.warning("[COUNTRY] %d ligne(s) sans OPERATING_UNIT_CODE reconnu "
                    "(flux=%s) → fichier 'autre'", len(groups[AUTRE]), flux_id)

    os.makedirs(output_dir, exist_ok=True)
    paths: dict[str, str] = {}
    for country, rows in sorted(groups.items()):
        fname = f"rapport_{_FILENAME_BY_COUNTRY.get(country, country.lower())}_{date_str}.xlsx"
        fpath = os.path.join(output_dir, fname)
        try:
            export_detailed_excel(
                rows, subset_stats(rows), f"{flux_id}|{COUNTRY_LABELS.get(country, AUTRE_LABEL)}",
                output_path=fpath,
            )
            paths[country] = fpath
            log.info("[COUNTRY] %s : %d lignes → %s", country, len(rows), fname)
        except Exception as e:
            log.warning("[COUNTRY] Échec export %s (non bloquant): %s", fname, e)

    return paths
