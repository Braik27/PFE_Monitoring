"""
engine/detailed_report.py
Rapport de comparaison détaillé "niveau consultant" pour le flux ITEMS.

Pour chaque article comparé, fournit :
  - le numéro de ligne exact dans le fichier CSV Cegid d'origine
  - le numéro de ligne exact dans le fichier CSV Oracle d'origine
  - les valeurs des deux côtés pour les colonnes comparées
  - un statut (OK / ECART / ABSENT_CEGID / ABSENT_ORACLE)
  - une explication en français prête à l'emploi
"""
from __future__ import annotations

import json
import logging
import re
from datetime import datetime
from typing import List, Optional

import pandas as pd

log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Construction du rapport détaillé
# ─────────────────────────────────────────────────────────────────────────────

def _resolve_col(merged: pd.DataFrame, base: str, side: str) -> Optional[str]:
    """Retourne le nom réel d'une colonne dans le merged DataFrame
    (gère les suffixes _cegid / _oracle)."""
    suffixed = f"{base}_{side}"
    if suffixed in merged.columns:
        return suffixed
    if base in merged.columns:
        return base
    return None


def _valeurs_egales(val_c, val_o, tolerance: float = 0.01) -> bool:
    """Compare deux valeurs en gérant NaN, 0, formats numériques et texte."""
    nan_c = pd.isna(val_c)
    nan_o = pd.isna(val_o)

    # Les deux NaN → égaux
    if nan_c and nan_o:
        return True

    str_c = str(val_c).strip() if not nan_c else ""
    str_o = str(val_o).strip() if not nan_o else ""

    # Vide/NaN vs 0 → égaux
    if (str_c in ("", "nan", "None", "none") and str_o in ("0", "0.0", "0.00", "")):
        return True
    if (str_o in ("", "nan", "None", "none") and str_c in ("0", "0.0", "0.00", "")):
        return True

    # Comparaison numérique avec tolérance
    try:
        num_c = float(str_c.replace(",", "."))
        num_o = float(str_o.replace(",", "."))
        if abs(num_c - num_o) <= tolerance:
            return True
    except (ValueError, TypeError):
        pass

    # Comparaison texte normalisée
    norm_c = re.sub(r"\s+", " ", str_c.upper())
    norm_o = re.sub(r"\s+", " ", str_o.upper())
    return norm_c == norm_o


def build_detail_report(
    merged: pd.DataFrame,
    cles: List[str],
    valeurs: List[str],
    flux_id: str,
    *,
    explanation_provider=None,
    comparison_rules: Optional[List[dict]] = None,
) -> List[dict]:
    """
    Construit le rapport détaillé "niveau consultant" à partir du DataFrame
    merged (résultat outer merge de comparer_flux).

    Args:
        merged:  DataFrame issu du merge outer (contient _merge, _LIGNE_FICHIER_*)
        cles:    Liste des colonnes clés utilisées pour le rapprochement
        valeurs: Liste des colonnes de valeur comparées
        flux_id: Identifiant du flux
        explanation_provider: callable(optionnel) — fonction
            (type_ecart, colonne, val_cegid, val_oracle) → str pour
            fournir une explication métier personnalisée.
            Si None, génère une explication par défaut.
        comparison_rules: list of dicts optionnel — règles du registry
            [{"column": "UNIT_PRICE", "severity": "CRITIQUE", "tolerance": 0.01}, ...]
            Utilisé pour déterminer la SEVERITE de chaque ligne.

    Returns:
        Liste de dicts, un par ligne de l'analyse détaillée.
    """
    if merged is None or merged.empty:
        return []

    # Construire lookup severity depuis comparison_rules
    severity_map: dict = {}
    if comparison_rules:
        for rule in comparison_rules:
            col = rule.get("column", "").upper() if isinstance(rule, dict) else getattr(rule, "column", "").upper()
            sev = rule.get("severity", "WARNING") if isinstance(rule, dict) else getattr(rule, "severity", "WARNING")
            if col:
                severity_map[col] = sev.upper()

    report: List[dict] = []

    # Résoudre les noms de colonnes (avec suffixes)
    col_ligne_cegid = _resolve_col(merged, "_LIGNE_FICHIER", "cegid")
    col_ligne_oracle = _resolve_col(merged, "_LIGNE_FICHIER", "oracle")

    # Colonnes de valeur à afficher (ex: UNIT_PRICE, ITEM_BARCODE, DESCRIPTION)
    display_cols = [v for v in valeurs if v.upper() not in {c.upper() for c in cles}]

    for _, row in merged.iterrows():
        merge_mark = row.get("_merge", "")

        # Clé de l'article
        key_val = " | ".join(str(row.get(c, "?")) for c in cles)

        # Numéros de ligne originaux
        ligne_cegid = row.get(col_ligne_cegid) if col_ligne_cegid else None
        ligne_oracle = row.get(col_ligne_oracle) if col_ligne_oracle else None

        # Convertir en int si possible, sinon laisser tel quel
        if pd.notna(ligne_cegid):
            try:
                ligne_cegid = int(ligne_cegid)
            except (ValueError, TypeError):
                ligne_cegid = str(ligne_cegid)
        else:
            ligne_cegid = None

        if pd.notna(ligne_oracle):
            try:
                ligne_oracle = int(ligne_oracle)
            except (ValueError, TypeError):
                ligne_oracle = str(ligne_oracle)
        else:
            ligne_oracle = None

        # Valeurs comparées
        valeurs_detail = {}
        for col in display_cols:
            val_c = row.get(f"{col}_cegid") if f"{col}_cegid" in merged.columns else row.get(col)
            val_o = row.get(f"{col}_oracle") if f"{col}_oracle" in merged.columns else row.get(col)
            valeurs_detail[col] = {
                "cegid":  str(val_c) if pd.notna(val_c) else "",
                "oracle": str(val_o) if pd.notna(val_o) else "",
            }

        # Statut et écarts détectés
        if merge_mark == "left_only":
            statut = "ABSENT_ORACLE"
            ecarts_detectes = []
            explication = "Article présent dans Cegid mais absent d'Oracle. Vérifier si la création article a été effectuée dans Oracle."
        elif merge_mark == "right_only":
            statut = "ABSENT_CEGID"
            ecarts_detectes = []
            explication = "Article présent dans Oracle mais absent de Cegid. Vérifier si l'article doit exister côté Cegid."
        else:
            # Comparaison des valeurs
            ecarts_detectes = []
            for col in display_cols:
                val_c = row.get(f"{col}_cegid") if f"{col}_cegid" in merged.columns else row.get(col)
                val_o = row.get(f"{col}_oracle") if f"{col}_oracle" in merged.columns else row.get(col)

                if _valeurs_egales(val_c, val_o):
                    continue

                ecarts_detectes.append(col)

            if ecarts_detectes:
                statut = "ECART"
            else:
                statut = "OK"

            # Explication
            if explanation_provider and ecarts_detectes:
                explication = explanation_provider(
                    ecarts_detectes[0] if len(ecarts_detectes) == 1 else "MULTIPLE",
                    ecarts_detectes,
                    {col: valeurs_detail.get(col, {}).get("cegid", "") for col in ecarts_detectes},
                    {col: valeurs_detail.get(col, {}).get("oracle", "") for col in ecarts_detectes},
                )
            elif ecarts_detectes:
                parts = []
                for col in ecarts_detectes:
                    vd = valeurs_detail.get(col, {})
                    parts.append(f"{col} : Cegid={vd.get('cegid', '?')} vs Oracle={vd.get('oracle', '?')}")
                explication = f"Écart détecté sur {', '.join(parts)}"
            else:
                explication = "Aucun écart — les valeurs sont conformes des deux côtés"

        # ── Déterminer la sévérité ──────────────────────────────────────
        if statut == "ABSENT_CEGID":
            severite = "CRITIQUE"
        elif statut == "ABSENT_ORACLE":
            severite = "WARNING"
        elif statut == "ECART" and ecarts_detectes:
            # Prendre la sévérité la plus grave parmi les colonnes en écart
            worst = "WARNING"
            for col in ecarts_detectes:
                s = severity_map.get(col.upper(), "WARNING")
                if s == "CRITIQUE":
                    worst = "CRITIQUE"
                    break
            severite = worst
        else:
            severite = ""

        # Construire la ligne du rapport
        report_row = {
            "ITEM_CODE":    key_val,
            "LIGNE_CEGID":  ligne_cegid,
            "LIGNE_ORACLE": ligne_oracle,
            "STATUT":       statut,
            "SEVERITE":     severite,
            "EXPLICATION":  explication,
        }

        # Ajouter les colonnes de valeur
        for col in display_cols:
            vd = valeurs_detail.get(col, {})
            report_row[f"{col}_CEGID"]  = vd.get("cegid", "")
            report_row[f"{col}_ORACLE"] = vd.get("oracle", "")

        report.append(report_row)

    # Trier par clé
    if cles and report:
        report.sort(key=lambda r: r.get("ITEM_CODE", ""))

    n_ok = sum(1 for r in report if r["STATUT"] == "OK")
    n_ecarts = sum(1 for r in report if r["STATUT"] == "ECART")
    n_abs_c = sum(1 for r in report if r["STATUT"] == "ABSENT_CEGID")
    n_abs_o = sum(1 for r in report if r["STATUT"] == "ABSENT_ORACLE")
    log.info(
        "[DETAIL] flux=%s — %d lignes : %d OK, %d écarts, %d absent Cegid, %d absent Oracle",
        flux_id, len(report), n_ok, n_ecarts, n_abs_c, n_abs_o,
    )
    return report


# ─────────────────────────────────────────────────────────────────────────────
# Export Excel 3 onglets
# ─────────────────────────────────────────────────────────────────────────────

# Codes couleur Excel
_FILL_OK       = "solid"   # vert pâle
_FG_OK         = "FFE8F5E9"
_FILL_ECART    = "solid"   # orange pâle
_FG_ECART      = "FFFFF3CD"
_FILL_ABSENT   = "solid"   # rouge pâle
_FG_ABSENT     = "FFFFE0E0"
_FILL_HEADER   = "solid"
_FG_HEADER     = "FF1F3864"


def export_detailed_excel(
    report: List[dict],
    stats: dict,
    flux_id: str,
    *,
    output_path: Optional[str] = None,
) -> str:
    """
    Exporte le rapport détaillé en Excel (.xlsx) avec 3 onglets :
      1. Résumé         — KPIs et statistiques globales
      2. Rapport détaillé — toutes les lignes, code couleur par statut
      3. Écarts uniquement — filtré sur lignes en écart ou absentes

    Args:
        report:      Liste de dicts retournée par build_detail_report()
        stats:       Dict de stats (nb_lignes_cegid, nb_lignes_oracle, etc.)
        flux_id:     Identifiant du flux
        output_path: Chemin de sortie (auto-généré si None)

    Returns:
        Chemin du fichier Excel créé.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    if output_path is None:
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        output_path = f"reports/rapport_detaille_{flux_id}_{ts}.xlsx"

    # S'assurer que le dossier existe
    import os
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb = Workbook()

    # Styles
    hdr_fill   = PatternFill(_FILL_HEADER, fgColor=_FG_HEADER)
    hdr_font   = Font(bold=True, color="FFFFFFFF", size=10, name="Calibri")
    cell_font  = Font(size=10, name="Calibri")
    title_font = Font(bold=True, size=14, name="Calibri")
    bold_font  = Font(bold=True, size=10, name="Calibri")

    fill_ok      = PatternFill(_FILL_OK,    fgColor=_FG_OK)
    fill_ecart   = PatternFill(_FILL_ECART, fgColor=_FG_ECART)
    fill_absent  = PatternFill(_FILL_ABSENT, fgColor=_FG_ABSENT)

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ════════════════════════════════════════════════════════════════════════
    # ONGLET 1 : Résumé
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Résumé"
    ws.sheet_properties.tabColor = "1F3864"

    ws.cell(row=1, column=1, value="Rapport détaillé — Comparaison Cegid vs Oracle").font = title_font
    ws.merge_cells("A1:D1")

    n_ok     = sum(1 for r in report if r["STATUT"] == "OK")
    n_ecarts = sum(1 for r in report if r["STATUT"] == "ECART")
    n_abs_c  = sum(1 for r in report if r["STATUT"] == "ABSENT_CEGID")
    n_abs_o  = sum(1 for r in report if r["STATUT"] == "ABSENT_ORACLE")
    n_critiques = sum(1 for r in report if r.get("SEVERITE") == "CRITIQUE")
    n_warnings  = sum(1 for r in report if r.get("SEVERITE") == "WARNING")
    total    = len(report)

    labels = [
        ("Flux",                  flux_id),
        ("Date du rapport",       datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ("",                      ""),
        ("Lignes Cegid (brut)",   stats.get("nb_lignes_cegid", "?")),
        ("Lignes Oracle (brut)",  stats.get("nb_lignes_oracle", "?")),
        ("",                      ""),
        ("Total articles analysés", total),
        ("Conformes (OK)",        n_ok),
        ("Écarts de valeur",      n_ecarts),
        ("Absents d'Oracle",      n_abs_o),
        ("Absents de Cegid",      n_abs_c),
        ("",                      ""),
        ("Taux de conformité",    f"{round(n_ok / total * 100, 1) if total else 0}%"),
        ("Écarts critiques",      n_critiques),
        ("Écarts warnings",       n_warnings),
    ]

    for i, (k, v) in enumerate(labels, 3):
        ws.cell(row=i, column=1, value=k).font = bold_font
        ws.cell(row=i, column=2, value=str(v)).font = cell_font

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    # ════════════════════════════════════════════════════════════════════════
    # ONGLET 2 : Rapport détaillé (toutes les lignes)
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Rapport détaillé")
    ws2.sheet_properties.tabColor = "3b82f6"

    # En-têtes
    headers = ["ITEM_CODE", "LIGNE_CEGID", "LIGNE_ORACLE", "STATUT", "SEVERITE"]
    # Ajouter les colonnes de valeur (CEGID + ORACLE)
    display_cols_seen = set()
    for r in report:
        for k in r:
            if k.endswith("_CEGID") and k not in headers:
                base = k[:-6]  # enlever _CEGID
                if base not in display_cols_seen:
                    display_cols_seen.add(base)
                    headers.append(f"{base}_CEGID")
                    headers.append(f"{base}_ORACLE")
    headers.append("EXPLICATION")

    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Données
    for r_idx, row_data in enumerate(report, 2):
        statut = row_data.get("STATUT", "")
        if statut == "OK":
            r_fill = fill_ok
        elif statut == "ECART":
            r_fill = fill_ecart
        else:
            r_fill = fill_absent

        for c_idx, h in enumerate(headers, 1):
            val = row_data.get(h, "")
            cell = ws2.cell(row=r_idx, column=c_idx, value=val if val is not None else "")
            cell.font = cell_font
            cell.fill = r_fill
            cell.border = thin_border

    # Largeurs de colonnes
    col_widths = {
        "ITEM_CODE": 18, "LIGNE_CEGID": 12, "LIGNE_ORACLE": 12,
        "STATUT": 16, "SEVERITE": 12, "EXPLICATION": 60,
    }
    for c_idx, h in enumerate(headers, 1):
        letter = chr(64 + c_idx) if c_idx <= 26 else chr(64 + (c_idx - 1) // 26) + chr(65 + (c_idx - 1) % 26)
        ws2.column_dimensions[letter].width = col_widths.get(h, 20)

    # Filtre automatique
    ws2.auto_filter.ref = f"A1:{chr(64 + len(headers))}1"

    # ════════════════════════════════════════════════════════════════════════
    # ONGLET 3 : Écarts uniquement
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Écarts uniquement")
    ws3.sheet_properties.tabColor = "e74c3c"

    ecarts_only = [r for r in report if r["STATUT"] != "OK"]

    for c, h in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for r_idx, row_data in enumerate(ecarts_only, 2):
        statut = row_data.get("STATUT", "")
        r_fill = fill_ecart if statut == "ECART" else fill_absent

        for c_idx, h in enumerate(headers, 1):
            val = row_data.get(h, "")
            cell = ws3.cell(row=r_idx, column=c_idx, value=val if val is not None else "")
            cell.font = cell_font
            cell.fill = r_fill
            cell.border = thin_border

    for c_idx, h in enumerate(headers, 1):
        letter = chr(64 + c_idx) if c_idx <= 26 else chr(64 + (c_idx - 1) // 26) + chr(65 + (c_idx - 1) % 26)
        ws3.column_dimensions[letter].width = col_widths.get(h, 20)

    if ecarts_only:
        ws3.auto_filter.ref = f"A1:{chr(64 + len(headers))}1"

    # Sauvegarder
    wb.save(output_path)
    log.info(
        "[DETAIL] Excel exporté : %s (%d lignes, %d écarts)",
        output_path, total, len(ecarts_only),
    )
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# Rapport mono-alerte pour email de breach
# ─────────────────────────────────────────────────────────────────────────────

def build_single_alert_report(alert: dict, anomalies: list = None) -> dict:
    """
    Build a compact payload for a single-alert breach email report.
    
    Returns dict with keys: flux_id, flux_name, concordance, severity,
    sla_status, workflow_status, anomalies (list), created_at.
    """
    anomalies = anomalies or []
    try:
        parsed = json.loads(alert.get("anomalies_json", "[]")) if alert.get("anomalies_json") else []
        anomalies = parsed if isinstance(parsed, list) else []
    except Exception:
        pass

    concordance = alert.get("concordance", 100.0)
    if concordance is not None:
        try:
            concordance = round(float(concordance), 1)
        except (TypeError, ValueError):
            concordance = 100.0

    return {
        "flux_id": alert.get("flux_id", ""),
        "flux_name": alert.get("flux_name", ""),
        "concordance": concordance,
        "severity": alert.get("severity") or alert.get("severity_class", ""),
        "sla_status": alert.get("sla_status", "ON_TIME"),
        "workflow_status": alert.get("workflow_status") or alert.get("status", "NEW"),
        "n_critiques": alert.get("n_critiques", 0),
        "n_warnings": alert.get("n_warnings", 0),
        "anomalies": anomalies,
        "created_at": alert.get("created_at", ""),
        "token": alert.get("token", ""),
    }
