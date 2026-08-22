import json, gzip
"""api/analysis.py — Lancement d'analyses + stats + reporting."""
import logging
import tempfile, os
import copy
from datetime import datetime, timedelta
import io
from flask import Blueprint, jsonify, request, session, Response, send_file
from engine.pipeline import AnalysisRequest, run_analysis
from core.email_alert import send_alert_async
from storage import get_storage
from api.auth import require_auth
from storage.base import json_encode

MAX_FILE_SIZE_MB = 20   # Limite à 20 Mégaoctets
# Nombre max d'anomalies sérialisées en DB.
# Au-delà, les compteurs (n_critiques etc.) sont exacts mais
# la liste détaillée est tronquée → évite timeout gunicorn sur Azure.
MAX_ANOMALIES_STORED = 1000

analysis_bp   = Blueprint("analysis", __name__)
UPLOAD_FOLDER = os.environ.get(
    "UPLOAD_FOLDER",
    os.path.join(tempfile.gettempdir(), "flux_uploads")
)

log = logging.getLogger(__name__)

DIVISIONS = ["KSA", "KWT", "SPG", "DAW7A", "LUX", "FRT", "LNH", "PSC", "ABA", "GLOBAL"]


def _summary_for_local_storage(summary: dict, blob_path: str | None = None) -> dict:
    """Keep local DB light while preserving exact counters and Azure pointer."""
    light = copy.deepcopy(summary)
    if blob_path:
        light["blob_path"]        = blob_path
        light["details_storage"]  = "azure_blob"

    for pair in light.get("pairs", []):
        anoms = pair.get("anomalies", [])
        pair["anomalies_total"]     = pair.get("anomalies_total", len(anoms))
        pair["anomalies_truncated"] = len(anoms) > 0
        pair["anomalies_in_db"]     = False
        pair["anomalies_in_blob"]   = bool(blob_path)
        if anoms and not pair.get("top_error_columns"):
            col_counts: dict[str, int] = {}
            for a in anoms:
                col = a.get("column") if isinstance(a, dict) else getattr(a, "column", None)
                if col:
                    col_counts[col] = col_counts.get(col, 0) + 1
            pair["top_error_columns"] = [
                {"column": c, "n_errors": n}
                for c, n in sorted(col_counts.items(), key=lambda x: -x[1])
            ]
        pair["anomalies"]           = []

    return light


def _backfill_top_error_columns_from_blob(row: dict, summary: dict) -> None:
    """One-time backfill: load full report, compute top_error_columns,
    and persist back to SQLite so subsequent reads are fast."""
    blob_path = summary.get("blob_path")
    if not blob_path:
        return
    try:
        from storage.blob_upload import download_report
        full_summary = download_report(blob_path)
        if not full_summary:
            log.warning("[BACKFILL] Empty or unreadable report for %s", blob_path)
            return
    except Exception as e:
        log.warning("[BACKFILL] Cannot read blob %s: %s", blob_path, e)
        return

    blob_pairs = full_summary.get("pairs", [])
    changed = False
    for pair in summary.get("pairs", []):
        if pair.get("top_error_columns"):
            continue
        for bp in blob_pairs:
            if bp.get("label") == pair.get("label") or bp.get("flux_id") == pair.get("flux_id"):
                anoms = bp.get("anomalies", [])
                if anoms:
                    col_counts: dict[str, int] = {}
                    for a in anoms:
                        col = a.get("column") if isinstance(a, dict) else None
                        if col:
                            col_counts[col] = col_counts.get(col, 0) + 1
                    pair["top_error_columns"] = [
                        {"column": c, "n_errors": n}
                        for c, n in sorted(col_counts.items(), key=lambda x: -x[1])
                    ]
                    changed = True
                break

    if changed:
        try:
            get_storage().update_summary(row["id"], summary)
        except Exception as e:
            log.warning("[BACKFILL] Failed to persist backfill for id=%s: %s", row.get("id"), e)


@analysis_bp.post("/api/analyze")
@require_auth
def analyze():
    flux_id  = request.form.get("flux_id", "").upper().strip()
    label    = request.form.get("label", "Analyse sans titre").strip()
    division = request.form.get("division", "").upper().strip()

    if not flux_id:
        return jsonify({"error": "flux_id est requis"}), 400

    f_cegid  = request.files.get("cegid")
    f_oracle = request.files.get("oracle")
    if not f_cegid or not f_oracle:
        return jsonify({"error": "Les fichiers cegid et oracle sont requis"}), 400

    # Vérifier la taille des fichiers
    f_cegid.seek(0, 2)
    taille_cegid = f_cegid.tell()
    f_cegid.seek(0)

    f_oracle.seek(0, 2)
    taille_oracle = f_oracle.tell()
    f_oracle.seek(0)

    taille_cegid_MB  = taille_cegid  / 1024 / 1024
    taille_oracle_MB = taille_oracle / 1024 / 1024

    log.info("Taille fichier Cegid: %.2f MB", taille_cegid_MB)
    log.info("Taille fichier Oracle: %.2f MB", taille_oracle_MB)

    if taille_cegid_MB > MAX_FILE_SIZE_MB or taille_oracle_MB > MAX_FILE_SIZE_MB:
        return jsonify({
            "error": f"Fichier trop volumineux. Maximum {MAX_FILE_SIZE_MB} MB"
        }), 400

    # Intègre la division dans le label si fournie
    if division and division not in label.upper():
        label = f"{label} [{division}]"

    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    with tempfile.NamedTemporaryFile(
            delete=False, dir=UPLOAD_FOLDER,
            suffix=os.path.splitext(f_cegid.filename)[1]) as fc:
        f_cegid.save(fc.name)
        path_cegid = fc.name

    with tempfile.NamedTemporaryFile(
            delete=False, dir=UPLOAD_FOLDER,
            suffix=os.path.splitext(f_oracle.filename)[1]) as fo:
        f_oracle.save(fo.name)
        path_oracle = fo.name

    try:
        req = AnalysisRequest(
            flux_id=flux_id,
            label=label,
            pairs=[{
                "cegid":  path_cegid,
                "oracle": path_oracle,
                "label":  f"{f_cegid.filename} vs {f_oracle.filename}",
            }],
            forced_division=division,
        )

        import time
        t0 = time.time()
        result = run_analysis(req)
        elapsed = time.time() - t0
        log.info("run_analysis %s terminé en %.1fs", flux_id, elapsed)
       
       
    except Exception as e:
        log.exception("Erreur lors de l'analyse du flux %s", flux_id)
        return jsonify({"error": f"Erreur serveur pendant l'analyse : {str(e)}"}), 500
    finally:
        for p in (path_cegid, path_oracle):
            try:
                os.unlink(p)
            except Exception:
                pass

    if result.error:
        return jsonify({"error": result.error}), 500

    # Résumé enrichi
    summary = result.to_dict()
    summary["division"]        = (result.divisions_found[0]
                                   if result.divisions_found else division or "GLOBAL")
    summary["divisions_found"] = result.divisions_found
    summary["analyst"]         = session.get("user", {}).get("username", "")

    # ── 1. Upload Blob avec anomalies complètes ──────────────────────────────────
    from storage.blob_upload import upload_report_to_blob

    summary_for_blob = copy.deepcopy(summary)
    for pair in summary_for_blob.get("pairs", []):
        pair.pop("anomalies_truncated", None)
        pair.pop("anomalies_in_db",     None)
        pair.pop("anomalies_in_blob",   None)
        pair["anomalies_total"] = len(pair.get("anomalies", []))

    blob_path = upload_report_to_blob(summary_for_blob, flux_id=result.flux_id)

    # ── 2. Version allégée pour SQLite (sans anomalies) ─────────────────────────
    local_summary = _summary_for_local_storage(summary, blob_path)

    # Sauvegarder dans le storage
    try:
        analysis_id = get_storage().save_analysis(
            flux_id=result.flux_id, label=label, summary=local_summary)
        send_alert_async(result, analysis_id=analysis_id)
    except Exception as e:
        log.exception("Erreur lors de la sauvegarde de l'analyse")
        return jsonify({"error": f"Erreur lors de la sauvegarde : {str(e)}"}), 500

    local_summary["analysis_id"] = analysis_id
    return _safe_json_response(local_summary)


@analysis_bp.get("/api/history")
@require_auth
def history():
    flux_id  = request.args.get("flux_id")
    division = request.args.get("division", "").upper().strip()
    limit    = int(request.args.get("limit", 50))
    rows     = get_storage().list_analyses(flux_id=flux_id, limit=limit)
    if division:
        rows = [r for r in rows if division in (r.get("label", "") or "").upper()]
    return jsonify(rows)


@analysis_bp.get("/api/analysis/<int:aid>/anomalies")
@require_auth
def get_anomalies(aid: int):
    """
    Retourne les anomalies paginées d'une analyse depuis Azure Blob.

    Query params:
        page     : numéro de page (défaut: 1)
        per_page : anomalies par page (défaut: 100, max: 1000)
        type     : filtre sur error_type (ex: MANQUANT_ORACLE)
        severity : filtre sur severity (ex: CRITIQUE)
    """
    row = get_storage().get_analysis(aid)
    if not row:
        return jsonify({"error": "Analyse introuvable"}), 404

    summary = row.get("summary", {}) or {}
    blob_path = summary.get("blob_path")

    if not blob_path:
        anomalies = []
        for pair in summary.get("pairs", []):
            anomalies.extend(pair.get("anomalies") or [])
        return jsonify({
            "anomalies": anomalies,
            "total":     len(anomalies),
            "page":      1,
            "per_page":  len(anomalies) or 1,
            "pages":     1,
            "blob_path": None,
            "source":    "local",
        })

    try:
        from storage.blob_upload import download_report
        full_summary = download_report(blob_path)
        if not full_summary:
            return jsonify({"error": "Impossible de charger le rapport d'analyse"}), 502
    except Exception as e:
        log.exception("[ANOMALIES] Erreur lecture Blob %s", blob_path)
        return jsonify({"error": f"Impossible de lire le rapport : {e}"}), 502

    all_anomalies = []
    for pair in full_summary.get("pairs", []):
        all_anomalies.extend(pair.get("anomalies") or [])

    error_type = request.args.get("type", "").strip().upper()
    severity   = request.args.get("severity", "").strip().upper()

    if error_type:
        all_anomalies = [a for a in all_anomalies
                         if a.get("error_type", "").upper() == error_type]
    if severity:
        all_anomalies = [a for a in all_anomalies
                         if a.get("severity", "").upper() == severity]

    total = len(all_anomalies)

    try:
        page     = max(1, int(request.args.get("page", 1)))
        per_page = min(1000, max(1, int(request.args.get("per_page", 100))))
    except ValueError:
        return jsonify({"error": "Paramètres page/per_page invalides"}), 400

    pages = max(1, -(-total // per_page))
    start = (page - 1) * per_page
    end   = start + per_page

    return _safe_json_response({
        "anomalies": all_anomalies[start:end],
        "total":     total,
        "page":      page,
        "per_page":  per_page,
        "pages":     pages,
        "blob_path": blob_path,
        "source":    "azure_blob",
    })


@analysis_bp.get("/api/analysis/<int:aid>/export/excel")
@require_auth
def export_analysis_excel(aid: int):
    """Export Excel d'une analyse (résumé + anomalies)."""
    row = get_storage().get_analysis(aid)
    if not row:
        return jsonify({"error": "Analyse introuvable"}), 404

    summary = row.get("summary", {}) or {}

    # ── Servir le rapport détaillé Excel s'il existe ────────────────────────
    detailed_path = summary.get("detailed_excel_path")
    if detailed_path and os.path.exists(detailed_path):
        created_at = row.get('created_at', '')
        date_str = created_at.strftime('%Y-%m-%d') if hasattr(created_at, 'strftime') else str(created_at)[:10]
        fname = f"rapport_detaille_{row.get('flux_id', 'unknown')}_{aid}_{date_str}.xlsx"
        return send_file(
            detailed_path, as_attachment=True, download_name=fname,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # ── Fallback : générer l'ancien format (résumé + anomalies) ─────────────
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        return jsonify({"error": "openpyxl non installé — pip install openpyxl"}), 500

    blob_path = summary.get("blob_path")

    # Récupérer les anomalies (Blob si dispo, sinon locales)
    all_anomalies = []
    if blob_path:
        try:
            from storage.blob_upload import download_report
            full = download_report(blob_path)
            if full:
                for pair in full.get("pairs", []):
                    all_anomalies.extend(pair.get("anomalies") or [])
                log.info("[EXPORT] anomalies lues : %d", len(all_anomalies))
            else:
                log.warning("[EXPORT] Impossible de charger le rapport depuis le chemin %s", blob_path)
        except Exception as e:
            log.error("[EXPORT] Erreur lors de la lecture du rapport : %s", e)
    else:
        log.warning("[EXPORT] blob_path absent en base")
    if not all_anomalies:
        for pair in summary.get("pairs", []):
            all_anomalies.extend(pair.get("anomalies") or [])

    if not all_anomalies:
        try:
            flux_id = row.get("flux_id", "")
            alerts  = get_storage().list_alerts(flux_id=flux_id, limit=10)
            for alert in alerts:
                if alert.get("analysis_id") == aid:
                    raw = alert.get("anomalies")
                    if isinstance(raw, list):
                        all_anomalies = raw
                    break
        except Exception as e:
            log.warning("[EXPORT] Fallback alerts échoué : %s", e)

    wb = Workbook()

    
    ws = wb.active
    ws.title = "Résumé"

    hdr_fill = PatternFill("solid", fgColor="FF1F3864")
    hdr_font = Font(bold=True, color="FFFFFFFF", size=10, name="Calibri")
    cell_font = Font(size=10, name="Calibri")

    ws.cell(row=1, column=1, value="Flux Monitor — Rapport d'analyse").font = Font(bold=True, size=14, name="Calibri")
    ws.merge_cells("A1:D1")

    labels = [
        ("ID", row.get("id")),
        ("Flux", row.get("flux_id")),
        ("Label", row.get("label")),
        ("Division", summary.get("division") or "GLOBAL"),
        ("Date", row.get("created_at")),
        ("Analyste", summary.get("analyst") or ""),
        ("Concordance moyenne", f"{summary.get('concordance_moyenne', 100)}%"),
        ("Total critiques", summary.get("total_critiques", 0)),
        ("Total warnings", summary.get("total_warnings", 0)),
        ("Total anomalies", summary.get("total_anomalies", 0)),
        ("Blob", blob_path or "N/A"),
    ]
    for i, (k, v) in enumerate(labels, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, size=10, name="Calibri")
        ws.cell(row=i, column=2, value=str(v) if v is not None else "").font = cell_font

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60

    # ── Sheet 2 : Anomalies ────────────────────────────────────────────
    ws2 = wb.create_sheet("Anomalies")
    headers = ["Type", "Sévérité", "Colonne", "Valeur Cegid", "Valeur Oracle",
               "Clé", "Ligne Cegid", "Ligne Oracle", "Explication", "Action"]
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"))

    fill_critique = PatternFill("solid", fgColor="FFFFE0E0")
    fill_warning  = PatternFill("solid", fgColor="FFFFF3CD")
    font_critique = Font(size=10, name="Calibri", color="FF8B0000")
    font_warning  = Font(size=10, name="Calibri", color="FF7B3F00")

    for r, a in enumerate(all_anomalies, 2):
        severity = a.get("severity", "").upper()
        row_fill = fill_critique if severity == "CRITIQUE" else (
                   fill_warning  if severity == "WARNING"  else None)
        row_font = font_critique if severity == "CRITIQUE" else (
                   font_warning  if severity == "WARNING"  else cell_font)

        col_name = a.get("column") or ""
        if not col_name:
            error_type = a.get("error_type", "")
            if error_type.startswith("ECART_"):
                col_name = error_type[len("ECART_"):]
            elif error_type.startswith("VALUE_NULLE_"):
                col_name = error_type[len("VALUE_NULLE_"):]

        values = [
            a.get("error_type", ""),
            a.get("severity", ""),
            col_name,
            a.get("val_cegid") or "",
            a.get("val_oracle") or "",
            a.get("key_str") or " | ".join(
                f"{k}={v}" for k, v in a.get("key_values", {}).items()),
            a.get("line_cegid") or "",
            a.get("line_oracle") or "",
            a.get("explication", ""),
            a.get("action", ""),
        ]
        for c, val in enumerate(values, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = row_font
            if row_fill:
                cell.fill = row_fill

    widths = [18, 12, 16, 18, 18, 30, 12, 12, 50, 40]
    for c, w in enumerate(widths, 1):
        ws2.column_dimensions[chr(64 + c)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    created_at = row.get('created_at', '')
    date_str = created_at.strftime('%Y-%m-%d') if hasattr(created_at, 'strftime') else str(created_at)[:10]
    fname = f"analyse_{row.get('flux_id', 'unknown')}_{aid}_{date_str}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@analysis_bp.get("/api/history/latest")
@require_auth
def history_latest():
    """
    Retourne la DERNIÈRE analyse de CHAQUE flux — version allégée.
    Utilisé par le Dashboard pour afficher une carte par flux sans
    télécharger des centaines d'analyses inutiles.

    Réponse : liste de { id, flux_id, label, created_at, summary: {...léger...} }
    Une seule entrée par flux_id (la plus récente).
    """
    from engine.flux_loader import FluxLoader

    flux_list = FluxLoader.list_all()
    result = []

    for cfg in flux_list:
        # On ne demande que la dernière analyse de CE flux (limit=1)
        rows = get_storage().list_analyses(flux_id=cfg.flux_id, limit=1)
        if not rows:
            continue

        r = rows[0]
        s = r.get("summary", {}) or {}

        # Allègement : on garde tout ce dont le Dashboard a besoin pour
        # afficher la carte (pairs avec leurs compteurs), mais SANS
        # le détail des anomalies — sauf les 15 premières pour l'aperçu
        # "Voir détails" qui en affiche 15 maximum (slice(0, 15) dans le frontend).
        needs_backfill = any(
            not pair.get("top_error_columns") for pair in s.get("pairs", [])
        )
        if needs_backfill and s.get("blob_path") and s.get("details_storage") == "azure_blob":
            _backfill_top_error_columns_from_blob(r, s)

        light_pairs = []
        for pair in s.get("pairs", []):
            light_pairs.append({
                "division":          pair.get("division"),
                "n_cegid":           pair.get("n_cegid"),
                "n_oracle":          pair.get("n_oracle"),
                "n_matched":         pair.get("n_matched"),
                "n_missing_oracle":  pair.get("n_missing_oracle"),
                "n_missing_cegid":   pair.get("n_missing_cegid"),
                "n_critiques":       pair.get("n_critiques"),
                "n_warnings":        pair.get("n_warnings"),
                "concordance":       pair.get("concordance"),
                "top_error_columns": pair.get("top_error_columns", []),
                # ⚠️ On garde SEULEMENT les 15 premières anomalies —
                # suffisant pour le bouton "Voir détails" qui n'en
                # affiche que 15 (slice(0, 15) côté React).
                "anomalies": (pair.get("anomalies") or [])[:15],
            })

        result.append({
            "id":         r.get("id"),
            "flux_id":    r.get("flux_id"),
            "label":      r.get("label"),
            "created_at": r.get("created_at"),
            "summary": {
                "flux_name":           s.get("flux_name"),
                "analyst":             s.get("analyst", ""),
                "division":            s.get("division", ""),
                "divisions_found":     s.get("divisions_found", []),
                "concordance_moyenne": s.get("concordance_moyenne", 100),
                "total_critiques":     s.get("total_critiques", 0),
                "total_warnings":      s.get("total_warnings", 0),
                "pairs":               light_pairs,
            }
        })

    return jsonify(result)


@analysis_bp.get("/api/history/<int:aid>")
@require_auth
def get_analysis(aid: int):
    row = get_storage().get_analysis(aid)
    if not row:
        return jsonify({"error": "Introuvable"}), 404
    return jsonify(row)


@analysis_bp.delete("/api/history/<int:aid>")
@require_auth
def delete_analysis(aid: int):
    get_storage().delete_analysis(aid)
    return jsonify({"ok": True})


#   @analysis_bp.get("/api/stats")
#   @require_auth
#   def stats():
#       """Stats globales par flux — utilisées par le dashboard."""
#       from engine.flux_loader import FluxLoader
#       flux_list = FluxLoader.list_all()
#       out = []
#       for cfg in flux_list:
#           rows         = get_storage().list_analyses(flux_id=cfg.flux_id, limit=500)
#           total_crit   = sum(r["summary"].get("total_critiques", 0) for r in rows)
#           total_warn   = sum(r["summary"].get("total_warnings",  0) for r in rows)
#           concordances = [r["summary"].get("concordance_moyenne", 100) for r in rows]
#           last = rows[0] if rows else None
#           out.append({
#               "flux_id":         cfg.flux_id,
#               "flux_name":       cfg.flux_name,
#               "color":           cfg.display.color,
#               "icon":            cfg.display.icon,
#               "direction":       cfg.direction,
#               "frequency":       cfg.frequency,
#               "objective":       cfg.objective,
#               "main_rule":       cfg.main_rule,
#               "key_columns":     cfg.key_columns,
#               "n_analyses":      len(rows),
#               "total_critiques": total_crit,
#               "total_warnings":  total_warn,
#               "concordance_moy": (round(sum(concordances) / len(concordances), 1)
#                                   if concordances else 100.0),
#               "last_analysis":   last,
#           })
#       return jsonify(out)
#
# REMPLACEZ-LE par ceci :
 
@analysis_bp.get("/api/stats")
@require_auth
def stats():
    """
    Stats globales par flux — utilisées par le dashboard.

    Version optimisée : on ne charge plus 500 analyses complètes par flux.
    count_analyses() pour le total, list_analyses(limit=1) pour la carte,
    et list_analyses(limit=20) pour les agrégations.
    """
    from engine.flux_loader import FluxLoader
    flux_list = FluxLoader.list_all()
    out = []

    for cfg in flux_list:
        n_analyses = get_storage().count_analyses(flux_id=cfg.flux_id)
        last_rows  = get_storage().list_analyses(flux_id=cfg.flux_id, limit=1)
        stat_rows  = get_storage().list_analyses(flux_id=cfg.flux_id, limit=20)

        total_crit   = 0
        total_warn   = 0
        concordances = []
        last_light   = None

        for r in stat_rows:
            s = r.get("summary", {}) or {}
            total_crit += s.get("total_critiques", 0)
            total_warn += s.get("total_warnings", 0)
            concordances.append(s.get("concordance_moyenne", 100))

        if last_rows:
            r = last_rows[0]
            s = r.get("summary", {}) or {}
            light_pairs = []
            for pair in s.get("pairs", []):
                light_pairs.append({
                    "division":          pair.get("division"),
                    "n_cegid":           pair.get("n_cegid"),
                    "n_oracle":          pair.get("n_oracle"),
                    "n_matched":         pair.get("n_matched"),
                    "n_missing_oracle":  pair.get("n_missing_oracle"),
                    "n_missing_cegid":   pair.get("n_missing_cegid"),
                    "n_critiques":       pair.get("n_critiques"),
                    "n_warnings":        pair.get("n_warnings"),
                    "concordance":       pair.get("concordance"),
                    "top_error_columns": pair.get("top_error_columns", []),
                    "anomalies": (pair.get("anomalies") or [])[:15],
                })
            last_light = {
                "id":         r.get("id"),
                "flux_id":    r.get("flux_id"),
                "label":      r.get("label"),
                "created_at": r.get("created_at"),
                "summary": {
                    "flux_name":           s.get("flux_name"),
                    "analyst":             s.get("analyst", ""),
                    "division":            s.get("division", ""),
                    "divisions_found":     s.get("divisions_found", []),
                    "concordance_moyenne": s.get("concordance_moyenne", 100),
                    "total_critiques":     s.get("total_critiques", 0),
                    "total_warnings":      s.get("total_warnings", 0),
                    "pairs":               light_pairs,
                }
            }

        out.append({
            "flux_id":         cfg.flux_id,
            "flux_name":       cfg.flux_name,
            "color":           cfg.display.color,
            "icon":            cfg.display.icon,
            "direction":       cfg.direction,
            "frequency":       cfg.frequency,
            "objective":       cfg.objective,
            "main_rule":       cfg.main_rule,
            "key_columns":     cfg.key_columns,
            "n_analyses":      n_analyses,
            "total_critiques": total_crit,
            "total_warnings":  total_warn,
            "concordance_moy": (round(sum(concordances) / len(concordances), 1)
                                if concordances else 100.0),
            "last_analysis":   last_light,
        })
    return jsonify(out)


@analysis_bp.get("/api/reporting")
@require_auth
def reporting():
    """Reporting détaillé par flux + division + période."""
    period   = request.args.get("period", "month")
    flux_id  = request.args.get("flux_id")
    division = request.args.get("division", "").upper().strip()
    now      = datetime.now()
    start    = _period_start(now, period)

    from engine.flux_loader import FluxLoader
    flux_list = FluxLoader.list_all()
    if flux_id:
        flux_list = [f for f in flux_list if f.flux_id == flux_id.upper()]

    result = []
    for cfg in flux_list:
        rows        = get_storage().list_analyses(flux_id=cfg.flux_id, limit=500)
        rows_period = [
            r for r in rows
            if _parse_date(r.get("created_at", "")) >= start
            and (not division or division in (r.get("label", "") or "").upper())
        ]

        tc       = sum(r["summary"].get("total_critiques", 0) for r in rows_period)
        tw       = sum(r["summary"].get("total_warnings",  0) for r in rows_period)
        concs    = [r["summary"].get("concordance_moyenne", 100) for r in rows_period]
        conc_moy = round(sum(concs) / len(concs), 1) if concs else 100.0

        div_stats = {}
        for r in rows_period:
            lbl = (r.get("label") or "").upper()
            from engine.division_splitter import _extract_division_from_label as _edl
            div = _edl(lbl) or "GLOBAL"
            if div not in div_stats:
                div_stats[div] = {"n": 0, "critiques": 0, "warnings": 0}
            div_stats[div]["n"]         += 1
            div_stats[div]["critiques"] += r["summary"].get("total_critiques", 0)
            div_stats[div]["warnings"]  += r["summary"].get("total_warnings",  0)

        col_counts = {}
        for r in rows_period:
            for pair in r["summary"].get("pairs", []):
                for top in pair.get("top_error_columns", []):
                    c = top["column"]
                    col_counts[c] = col_counts.get(c, 0) + top["n_errors"]
        top_cols = sorted(
            [{"column": c, "n_errors": n} for c, n in col_counts.items()],
            key=lambda x: -x["n_errors"]
        )[:5]

        timeline = _build_timeline(rows, days=30,   division_filter=division)
        weekly   = _build_weekly  (rows, weeks=4,   division_filter=division)

        result.append({
            "flux_id":           cfg.flux_id,
            "flux_name":         cfg.flux_name,
            "icon":              cfg.display.icon,
            "color":             cfg.display.color,
            "direction":         cfg.direction,
            "period":            period,
            "n_analyses":        len(rows_period),
            "total_critiques":   tc,
            "total_warnings":    tw,
            "concordance_moy":   conc_moy,
            "top_error_columns": top_cols,
            "timeline":          timeline,
            "weekly":            weekly,
            "div_stats":         div_stats,
        })
    return jsonify(result)


# ── Helpers ────────────────────────────────────────────────────────────────────

def _safe_json_response(data, status=200):
    """
    Sérialise data en JSON avec l'encoder robuste (gère np.int64, NaN, datetime).
    """
    return Response(
        json_encode(data),
        status=status,
        mimetype="application/json"
    )

def _period_start(now, period):
    if period == "week":
        return now - timedelta(days=7)
    if period == "year":
        return now.replace(month=1, day=1, hour=0, minute=0, second=0)
    return now.replace(day=1, hour=0, minute=0, second=0)


def _parse_date(s):
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:19], fmt)
        except Exception:
            continue
    return datetime.min


def _build_timeline(rows, days=30, division_filter=""):
    from collections import defaultdict
    grouped = defaultdict(lambda: {"n": 0, "critiques": 0, "warnings": 0})
    for r in rows:
        date = r.get("created_at", "")[:10]
        if not division_filter or division_filter in (r.get("label", "") or "").upper():
            g = grouped[date]
            g["n"] += 1
            g["critiques"] += r.get("summary", {}).get("total_critiques", 0)
            g["warnings"] += r.get("summary", {}).get("total_warnings", 0)
    now = datetime.now()
    result = []
    for i in range(days - 1, -1, -1):
        d  = (now - timedelta(days=i)).strftime("%Y-%m-%d")
        g  = grouped.get(d, {"n": 0, "critiques": 0, "warnings": 0})
        result.append({
            "date":  d,
            "label": (now - timedelta(days=i)).strftime("%d/%m"),
            **g,
        })
    return result


def _build_weekly(rows, weeks=4, division_filter=""):
    now = datetime.now()
    result = []
    for w in range(weeks - 1, -1, -1):
        ws = now - timedelta(days=now.weekday() + 7 * w)
        we = ws + timedelta(days=6)
        n = 0
        crit = 0
        warn = 0
        for r in rows:
            rd = r.get("created_at", "")[:10]
            if ws.strftime("%Y-%m-%d") <= rd <= we.strftime("%Y-%m-%d"):
                if not division_filter or division_filter in (r.get("label", "") or "").upper():
                    n += 1
                    crit += r.get("summary", {}).get("total_critiques", 0)
                    warn += r.get("summary", {}).get("total_warnings", 0)
        result.append({
            "label":     f"S{ws.strftime('%d/%m')}",
            "n":         n,
            "critiques": crit,
            "warnings":  warn,
        })
    return result

