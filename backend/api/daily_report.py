"""
api/daily_report.py — Rapport Excel journalier Flux Monitor

FORMAT EXACT identique au fichier client ABA_LUX_Monitoring.xlsx :
  Colonnes : Direction | Flow Name | IN/OUT Cegid | Staging | Oracle | Status | Comment
  Couleurs : OK=vert | TBC=orange | KO=rouge
  Structure : 1 onglet par jour nommé "JJ-MM-AAAA"

LOGIQUE PAYS (mapping client ABA — source de vérité utilisateur) :
   - L'analyse est GLOBALE (pas de split lors du calcul)
   - Le pays est déduit des codes stockés dans summary.division(s_found)
     via OU_COUNTRY_MAP / LEGACY_BUCKET_COUNTRY (voir tables ci-dessous)
   - Pour le rapport par pays : on filtre les ANALYSES (pas les lignes)
   - Une analyse multi-pays apparaît dans chaque fichier concerné
   - Code inconnu/GLOBAL → bucket "Autre / Non classé" (visible + loggé)

Routes :
   GET /api/report/daily              → rapport du jour (toutes analyses)
   GET /api/report/daily?division=KWT → rapport Koweït uniquement
   GET /api/report/by-division        → ZIP : 1 fichier Excel par pays
   GET /api/report/monthly            → rapport mensuel (1 onglet/jour)
   GET /api/report/divisions          → liste pays disponibles
"""
from __future__ import annotations
import io, zipfile, logging
from datetime import datetime, timedelta
from flask import Blueprint, jsonify, request, send_file
from api.auth import require_auth
from storage import get_storage

log = logging.getLogger(__name__)

report_bp = Blueprint("report", __name__)

# ── openpyxl ─────────────────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Couleurs (exactes du fichier client) ─────────────────────────────
C_OK     = "FF70AD47"   # vert
C_TBC    = "FFBE5014"   # orange
C_KO     = "FFFF0000"   # rouge
C_HEADER = "FF1F3864"   # bleu foncé en-tête colonnes
C_TITLE  = "FF2460E8"   # bleu titre ligne 1
C_ADF    = "FF2F5496"   # bleu ADF Execution
C_EXP_BG = "FFD6E4F0"   # fond ligne Export (Cegid→Oracle)
C_IMP_BG = "FFEDE7F6"   # fond ligne Import (Oracle→Cegid)
C_WHITE  = "FFFFFFFF"
C_BLACK  = "FF000000"
C_GREY   = "FF595959"

# ── Mapping pays ──────────────────────────────────────────────────────
# Source de vérité unique : engine/country_detail_report.py (mapping ABA
# validé par le client). Importé ici pour les filtres et libellés des
# rapports de synthèse.
from engine.country_detail_report import (  # noqa: E402
    OU_COUNTRY_MAP, COUNTRY_LABELS, AUTRE, AUTRE_LABEL,
    LEGACY_BUCKET_COUNTRY, _FILENAME_BY_COUNTRY,
)


def _analysis_day(a: dict) -> str:
    """
    Date AAAA-MM-JJ d'une analyse, tolérante au type de created_at :
    datetime.datetime / datetime.date (MySQL, stockage local) → strftime,
    str ISO (backend Azure) → tronqué à 10 caractères.
    Un slicing [:10] direct lève TypeError sur un datetime.
    """
    created = a.get("created_at", "")
    if hasattr(created, "strftime"):
        return created.strftime("%Y-%m-%d")
    return str(created or "")[:10]

def _resolve_country_param(raw: str) -> str:
    """
    Résout un paramètre ?division= en pays : accepte un pays
    (QATAR/KUWAIT/KSA) ou un ancien code de division/bucket.
    Valeur inconnue → AUTRE (filtre explicite sur le non-classé).
    """
    t = str(raw or "").strip().upper()
    if t in COUNTRY_LABELS or t == AUTRE:
        return t
    return OU_COUNTRY_MAP.get(t) or LEGACY_BUCKET_COUNTRY.get(t) or AUTRE


def analysis_countries(a: dict) -> set:
    """
    Pays d'une analyse, déduits des codes stockés dans son summary
    (division + divisions_found — sortie historique de la détection OU).
    Retourne {AUTRE} si rien de classable : jamais silencieusement perdue.
    """
    s = a.get("summary", {}) or {}
    countries: set = set()
    for raw in [s.get("division"), *(s.get("divisions_found") or [])]:
        token = str(raw or "").strip().upper()
        if not token or token == "GLOBAL":
            continue
        c = OU_COUNTRY_MAP.get(token) or LEGACY_BUCKET_COUNTRY.get(token)
        countries.add(c if c else AUTRE)
    return countries or {AUTRE}


# ── Helpers Excel ─────────────────────────────────────────────────────

def _f(c):
    return PatternFill("solid", fgColor=c)

def _font(bold=False, color=C_BLACK, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, name="Aptos Narrow", italic=italic)

def _align(h="center", v="center"):
    return Alignment(horizontal=h, vertical=v)

def _border():
    s = Side(style="thin", color="FFB0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)

def _setup_cols(ws):
    """Largeurs de colonnes simplifiées."""
    widths = {"A":18, "B":28, "C":12, "D":12, "E":10, "F":45}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def _write_title(ws, row_idx, text):
    """Ligne 1 : titre bleu fusionné A–F."""
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
    c = ws.cell(row=row_idx, column=1, value=text)
    c.font      = _font(bold=True, color=C_WHITE, size=11)
    c.fill      = _f(C_TITLE)
    c.alignment = _align(h="left")
    ws.row_dimensions[row_idx].height = 20

def _write_header(ws, row_idx, flux_id=""):
    """Ligne 2 : en-têtes colonnes, fond bleu foncé."""
    cols = ["Direction", "Flow Name", "IN Cegid", "OUT Oracle", "Status", "Comment"]
    if flux_id == "CUSTOMERBALANCE":
        cols = ["Direction", "Flow Name", "IN Cegid", "OUT Oracle", "Nb Integrated", "Nb Rejected", "Status", "Comment"]
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=row_idx, column=i, value=h)
        c.font      = _font(bold=True, color=C_WHITE, size=10)
        c.fill      = _f(C_HEADER)
        c.alignment = _align()
        c.border    = _border()
    ws.row_dimensions[row_idx].height = 17

def _write_data_row(ws, row_idx, direction, flow_name,
                    n_cegid, n_oracle, status, n_errors, comment,
                    flux_id="", n_integrated=0, n_rejected=0):
    """
    Écrit une ligne de données.
    Fond : bleu clair = Export (Cegid→Oracle) | violet clair = Import (Oracle→Cegid)
    Statut : vert=OK | orange=TBC | rouge=KO
    """
    is_export = str(direction).upper().startswith("CEGID")
    row_bg = C_EXP_BG if is_export else C_IMP_BG

    su = (status or "").upper()
    s_bg = (C_OK  if su == "OK"
            else C_TBC if su == "TBC"
            else C_KO  if su in ("KO","ERROR","ERREUR")
            else "FFFFFFFF")
    s_fc = C_BLACK if su == "OK" else C_WHITE

    # Si flux CustomerBalance, ajoute les colonnes Nb Integrated / Nb Rejected
    if flux_id == "CUSTOMERBALANCE":
        vals = [direction, flow_name, n_cegid, n_oracle, n_integrated, n_rejected, status, n_errors, comment]
    else:
        vals = [direction, flow_name, n_cegid, n_oracle, status, n_errors, comment]
    
    for i, val in enumerate(vals, 1):
        c = ws.cell(row=row_idx, column=i, value=val)
        c.border    = _border()
        # Alignement selon le nombre de colonnes
        if flux_id == "CUSTOMERBALANCE":
            c.alignment = _align(h="left" if i in (1, 2, 9) else "center")
        else:
            c.alignment = _align(h="left" if i in (1, 2, 7) else "center")
        
        if i <= 4:
            c.fill = _f(row_bg)
            c.font = _font(bold=(i in (3, 4)), size=10)
        elif i == 5 and flux_id == "CUSTOMERBALANCE":
            # Nb Integrated - toujours vert
            c.fill = _f(C_OK)
            c.font = _font(bold=True, color=C_BLACK, size=10)
        elif i == 6 and flux_id == "CUSTOMERBALANCE":
            # Nb Rejected - rouge si > 0, sinon vert
            err_bg = C_KO if n_rejected > 0 else C_OK
            err_fc = C_WHITE if n_rejected > 0 else C_BLACK
            c.fill = _f(err_bg)
            c.font = _font(bold=True, color=err_fc, size=10)
        elif (i == 5 and flux_id != "CUSTOMERBALANCE") or (i == 7 and flux_id == "CUSTOMERBALANCE"):
            c.fill = _f(s_bg)
            c.font = _font(bold=True, color=s_fc, size=10)
        elif (i == 6 and flux_id != "CUSTOMERBALANCE") or (i == 8 and flux_id == "CUSTOMERBALANCE"):
            err_bg = C_KO if isinstance(val, int) and val > 0 else row_bg
            err_fc = C_WHITE if isinstance(val, int) and val > 0 else C_BLACK
            c.fill = _f(err_bg)
            c.font = _font(bold=True, color=err_fc, size=10)
        else:
            c.font = _font(size=9, italic=True, color=C_GREY)
    ws.row_dimensions[row_idx].height = 15


def _write_legend_sheet(wb):
    """Onglet Legend simplifié."""
    ws = wb.create_sheet(title="Legend")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 55
    rows = [
        ("Colonne",        "Signification"),
        ("IN Cegid",       "Nombre de lignes dans le fichier Cegid"),
        ("OUT Oracle",     "Nombre de lignes dans le fichier Oracle"),
        ("",               ""),
        ("Status",         "Signification"),
        ("OK",             "Toutes les données correspondent"),
        ("TBC",            "To Be Confirmed — écarts mineurs à vérifier"),
        ("KO",             "Erreurs critiques — intervention requise immédiate"),
        ("Errors",         "Nombre total d'erreurs (critiques + warnings)"),
    ]
    status_bg = {"OK": C_OK, "TBC": C_TBC, "KO": C_KO}
    for i, (a, b) in enumerate(rows, 1):
        c1 = ws.cell(row=i, column=1, value=a)
        c2 = ws.cell(row=i, column=2, value=b)
        if i == 1 or a == "Status":
            for c in (c1, c2):
                c.font = _font(bold=True, color=C_WHITE)
                c.fill = _f(C_HEADER)
        elif a in status_bg:
            fc = C_BLACK if a == "OK" else C_WHITE
            c1.fill = _f(status_bg[a])
            c1.font = _font(bold=True, color=fc)
        for c in (c1, c2):
            c.border    = _border()
            c.alignment = _align(h="left")


# ── Construction d'un onglet journée ─────────────────────────────────

def _build_sheet(ws, day_label: str, analyses: list, subtitle: str = ""):
    """
    Construit un onglet complet pour une journée.
    Une ligne par analyse/flux.
    Structure : Titre → En-tête → 1 ligne par flux
    """
    # Determine if any analysis is for CustomerBalance to adjust columns
    has_customerbalance = any(
        "CUSTOMERBALANCE" in (a.get("flux_id", "") + " " + a.get("label", "")).upper()
        for a in analyses
    )

    _setup_cols(ws)
    ws.freeze_panes = "A3"

    r = 1
    title = f"  Flux Monitor — {day_label}" + (f" | {subtitle}" if subtitle else "")
    _write_title(ws, r, title)
    r += 1
    # Pass flux_id to header to show extra columns for CustomerBalance
    flux_id_header = "CUSTOMERBALANCE" if has_customerbalance else ""
    _write_header(ws, r, flux_id=flux_id_header)
    r += 1

    for a in analyses:
        s       = a.get("summary", {})
        pairs   = s.get("pairs", [])
        flux_id = a.get("flux_id", "")
        is_customerbalance = "CUSTOMERBALANCE" in (a.get("flux_id", "") + " " + a.get("label", "")).upper()

        from engine.flux_loader import FluxLoader
        try:
            cfg       = FluxLoader.load(flux_id)
            direction = "Oracle --> Cegid" if cfg.direction == "import" else "Cegid --> Oracle"
            flow_name = cfg.flux_name
        except Exception:
            direction = "Cegid --> Oracle"
            flow_name = flux_id

        for pair in pairs:
            n_cegid = pair.get("n_cegid")
            n_oracle= pair.get("n_oracle")
            
            n_crit = (pair.get("n_critiques", 0)
                       + pair.get("n_missing_oracle", 0)
                       + pair.get("n_missing_cegid", 0))
            n_warn = pair.get("n_warnings", 0)
            n_errors = n_crit + n_warn
            conc   = pair.get("concordance", 100)

            if n_crit > 0:
                status = "KO"
            elif n_warn > 0 or conc < 100:
                status = "TBC"
            else:
                status = "OK"

            comment = _build_comment(pair, n_crit, n_warn, conc, a.get("label",""))

            # Get integrated/rejected counts for CustomerBalance
            n_integrated = pair.get("n_integrated", 0) if pair.get("n_integrated") is not None else 0
            n_rejected = pair.get("n_rejected", 0) if pair.get("n_rejected") is not None else 0

            _write_data_row(ws, r, direction, flow_name,
                            n_cegid or "—",
                            n_oracle or "—",
                            status, n_errors, comment,
                            flux_id="CUSTOMERBALANCE" if is_customerbalance else "",
                            n_integrated=n_integrated,
                            n_rejected=n_rejected)
            r += 1


def _build_comment(pair: dict, n_crit: int, n_warn: int, conc: float, label: str) -> str:
    """
    Génère un commentaire métier clair.
    Priorité : erreurs critiques → warnings → OK
    PAS de codes internes comme ERREUR_LECTURE.
    """
    anomalies = pair.get("anomalies", [])

    # Cas erreur de lecture — message explicite
    read_errors = [a for a in anomalies if a.get("error_type") == "ERREUR_LECTURE"]
    if read_errors:
        msg = read_errors[0].get("explication", "Erreur de lecture du fichier")
        # Nettoie le message : enlève les codes Python et les chemins
        for bad in ["ImportError","ModuleNotFoundError","No module named","engine.division_splitter"]:
            if bad in msg:
                return "Erreur d'import interne — contacter l'équipe technique"
        return f"Erreur lecture : {msg[:120]}"

    if n_crit == 0 and n_warn == 0 and conc >= 100:
        return "OK – données cohérentes"

    parts = []

    # Résumé chiffré
    n_c = pair.get("n_cegid", 0) or 0
    n_o = pair.get("n_oracle", 0) or 0
    if n_c != n_o and (n_c > 0 or n_o > 0):
        parts.append(f"Écart lignes : Cegid={n_c} | Oracle={n_o}")

    miss_o = pair.get("n_missing_oracle", 0)
    miss_c = pair.get("n_missing_cegid", 0)
    if miss_o > 0:
        parts.append(f"{miss_o} ligne(s) absente(s) dans Oracle")
    if miss_c > 0:
        parts.append(f"{miss_c} ligne(s) absente(s) dans Cegid")

    if n_crit > 0:
        parts.append(f"{n_crit} erreur(s) critique(s)")
    if n_warn > 0:
        parts.append(f"{n_warn} warning(s)")

    # Top colonnes en erreur
    top_cols = pair.get("top_error_columns", [])
    if top_cols:
        col_names = ", ".join(c["column"] for c in top_cols[:3])
        parts.append(f"Colonnes : {col_names}")

    return " | ".join(parts) if parts else f"Concordance {conc}%"


# ── Routes ────────────────────────────────────────────────────────────

@report_bp.get("/api/report/daily")
@require_auth
def daily_report():
    """
    Rapport du jour pour UN pays ou toutes.
    ?date=2026-04-13   (défaut = aujourd'hui)
    ?division=KWT      (filtre — accepte pays QATAR/KUWAIT/KSA ou ancien code)
    ?flux_id=SALES     (filtre par flux)
    """
    if not OPENPYXL_OK:
        return jsonify({"error": "openpyxl non installé — pip install openpyxl"}), 500

    division = request.args.get("division", "").strip()
    flux_id  = request.args.get("flux_id", "").upper().strip()
    date_str = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))

    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return jsonify({"error": "Format date invalide — AAAA-MM-JJ"}), 400

    # Récupère les analyses du jour
    all_a = get_storage().list_analyses(flux_id=flux_id or None, limit=1000)
    day_a = [a for a in all_a if _analysis_day(a) == date_str]

    # Filtre par pays si demandé
    country = ""
    if division:
        country = _resolve_country_param(division)
        day_a = [a for a in day_a if country in analysis_countries(a)]

    if not day_a:
        return jsonify({
            "error": f"Aucune analyse pour le {date_str}"
                     + (f" ({COUNTRY_LABELS.get(country, AUTRE_LABEL)})" if country else "")
        }), 404

    wb = Workbook()
    wb.remove(wb.active)
    day_label = target_date.strftime("%d-%m-%Y")

    ws = wb.create_sheet(title=day_label)
    _build_sheet(ws, day_label, day_a,
                 subtitle=COUNTRY_LABELS.get(country, AUTRE_LABEL) if country else "")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"FluxMonitor_{date_str}{'_' + _FILENAME_BY_COUNTRY.get(country, country) if country else ''}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@report_bp.get("/api/report/by-division")
@require_auth
def report_by_division():
    """
    Rapport détaillé LIGNE par ligne, un fichier Excel par pays → ZIP.

    Deux modes :
      ?analysis_id=42   → les fichiers pays stockés pour CETTE analyse
                          (bouton "Rapport par division" de la page Analyse)
      ?date=2026-04-13  → fusion des lignes de TOUTES les analyses du jour,
                          groupées par pays (ex : Items + Sales le même jour)

    Chaque fichier = même structure que l'"Excel analyse"
    (Résumé / Rapport détaillé / Écarts), filtré sur les lignes dont
    l'OPERATING_UNIT_CODE mappe vers ce pays. Génère par exemple :
      rapport_qatar_2026-04-13.xlsx
      rapport_kuwait_2026-04-13.xlsx
      rapport_ksa_2026-04-13.xlsx
      rapport_autre_2026-04-13.xlsx   ← lignes sans code reconnu (à surveiller)

    Les fichiers sont générés à la fin de chaque analyse (worker) et stockés
    dans le summary (country_excel_paths). Les analyses antérieures à cette
    fonctionnalité n'en ont pas → erreur explicite.
    """
    if not OPENPYXL_OK:
        return jsonify({"error": "openpyxl non installé"}), 500

    storage = get_storage()

    # ── Mode 1 : une analyse précise ────────────────────────────────────────
    analysis_id = request.args.get("analysis_id", "").strip()
    if analysis_id:
        try:
            analysis_id = int(analysis_id)
        except ValueError:
            return jsonify({"error": "analysis_id invalide"}), 400

        rec = storage.get_analysis(analysis_id)
        if not rec:
            return jsonify({"error": f"Analyse {analysis_id} introuvable"}), 404

        paths_by_country: dict[str, str] = (rec.get("summary") or {}).get(
            "country_excel_paths") or {}
        if not paths_by_country:
            return jsonify({"error":
                "Aucun rapport détaillé par pays stocké pour cette analyse — "
                "les analyses antérieures à cette fonctionnalité ne peuvent pas "
                "être reconstruites. Relancez une analyse pour obtenir le détail "
                "par pays."}), 404

        return _send_country_files(paths_by_country,
                                   date_str=(rec.get("created_at") or "")[:10])

    # ── Mode 2 : toutes les analyses d'une journée ──────────────────────────
    date_str = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    flux_id  = request.args.get("flux_id", "").upper().strip()

    try:
        datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return jsonify({"error": "Format date invalide — AAAA-MM-JJ"}), 400

    day_a = [a for a in storage.list_analyses(flux_id=flux_id or None, limit=1000)
             if _analysis_day(a) == date_str]
    if not day_a:
        return jsonify({"error": f"Aucune analyse pour le {date_str}"}), 404

    # Fusionne les LIGNES de toutes les analyses du jour, par pays
    from engine.country_detail_report import collect_day_rows, subset_stats
    from engine.detailed_report import export_detailed_excel

    maps = [(a.get("id"), (a.get("summary") or {}).get("country_excel_paths") or {})
            for a in day_a]
    groups, skipped = collect_day_rows(maps)

    if skipped:
        log.warning(
            "[REPORT] %d/%d analyse(s) du %s sans fichiers pays exploitables "
            "(antérieures à la fonctionnalité ou fichiers absents) — ignorées : %s",
            len(skipped), len(day_a), date_str, skipped,
        )
    if not groups:
        return jsonify({"error":
            f"Aucune analyse du {date_str} ne possède de rapport détaillé par "
            "pays stocké. Relancez une analyse pour l'obtenir."}), 404

    import os as _os
    import tempfile as _tempfile

    n_src = len(day_a) - len(skipped)
    out_dir = _tempfile.mkdtemp(prefix=f"flux_pays_{date_str}_")
    paths: dict[str, str] = {}
    for country, rows in sorted(groups.items()):
        fname = f"rapport_{_FILENAME_BY_COUNTRY.get(country, country.lower())}_{date_str}.xlsx"
        fpath = _os.path.join(out_dir, fname)
        try:
            export_detailed_excel(
                rows, subset_stats(rows),
                f"{COUNTRY_LABELS.get(country, AUTRE_LABEL)} | {n_src} analyse(s)",
                output_path=fpath,
            )
            paths[country] = fpath
        except Exception as e:
            log.warning("[REPORT] Échec fusion %s (non bloquant): %s", fname, e)

    if not paths:
        return jsonify({"error": "Fichiers de rapport introuvables sur le serveur"}), 404

    return _send_country_files(paths, date_str=date_str)


def _send_country_files(paths_by_country: dict[str, str], *, date_str: str):
    """Zippe (ou envoie seul) les fichiers Excel pays existants sur disque."""
    import os

    available = {c: p for c, p in paths_by_country.items() if p and os.path.exists(p)}
    missing = sorted(set(paths_by_country) - set(available))
    if missing:
        log.warning("[REPORT] Fichiers pays manquants sur disque : %s", missing)
    if not available:
        return jsonify({"error": "Fichiers de rapport introuvables sur le serveur"}), 404

    if AUTRE in available:
        log.warning("[REPORT] Le rapport contient un fichier 'autre' "
                    "(lignes sans OPERATING_UNIT_CODE reconnu)")

    if len(available) == 1:
        c, p = next(iter(available.items()))
        return send_file(p, as_attachment=True,
                         download_name=os.path.basename(p),
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for c, p in sorted(available.items()):
            zf.write(p, os.path.basename(p))

    zip_buf.seek(0)
    return send_file(zip_buf, as_attachment=True,
                     download_name=f"FluxMonitor_Pays_{date_str}.zip",
                     mimetype="application/zip")


@report_bp.get("/api/report/monthly")
@require_auth
def monthly_report():
    """
    Rapport mensuel — un onglet par jour.
    ?month=2026-04
    ?division=KWT   (filtre — accepte pays QATAR/KUWAIT/KSA ou ancien code)
    ?flux_id=SALES  (filtre par flux)
    """
    if not OPENPYXL_OK:
        return jsonify({"error": "openpyxl non installé"}), 500

    month_str = request.args.get("month", datetime.now().strftime("%Y-%m"))
    division  = request.args.get("division", "").strip()
    flux_id   = request.args.get("flux_id", "").upper().strip()

    try:
        first_day = datetime.strptime(month_str + "-01", "%Y-%m-%d")
    except ValueError:
        return jsonify({"error": "Format mois invalide — AAAA-MM"}), 400

    all_a = get_storage().list_analyses(flux_id=flux_id or None, limit=3000)
    wb = Workbook(); wb.remove(wb.active)
    sheets_created = 0

    country = _resolve_country_param(division) if division else ""
    current = first_day
    while current.month == first_day.month and current <= datetime.now():
        date_str  = current.strftime("%Y-%m-%d")
        day_label = current.strftime("%d-%m-%Y")

        day_a = [a for a in all_a if _analysis_day(a) == date_str]
        if country:
            day_a = [a for a in day_a if country in analysis_countries(a)]

        if day_a:
            ws = wb.create_sheet(title=day_label)
            _build_sheet(ws, day_label, day_a)
            sheets_created += 1

        current += timedelta(days=1)

    if sheets_created == 0:
        return jsonify({"error": f"Aucune donnée pour {month_str}"}), 404

    _write_legend_sheet(wb)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = (f"FluxMonitor_{month_str}"
             + ('_' + _FILENAME_BY_COUNTRY.get(country, country.lower()) if country else '')
             + ".xlsx")
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@report_bp.get("/api/report/divisions")
@require_auth
def list_divisions():
    """Pays présents dans les analyses stockées (mapping ABA)."""
    analyses = get_storage().list_analyses(limit=1000)
    countries: set = set()
    for a in analyses:
        countries |= analysis_countries(a)
    return jsonify(sorted(COUNTRY_LABELS.get(c, AUTRE_LABEL) for c in countries))