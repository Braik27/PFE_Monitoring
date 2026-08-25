"""api/customerbalance_report.py - Rapport spécifique pour le flux CustomerBalance.
Gère les lignes Rejected (R) vs Integrated (I) du fichier Cegid.
"""
from __future__ import annotations
from flask import Blueprint, jsonify, request, make_response
from api.auth import require_auth
import io

cb_bp = Blueprint("customerbalance_report", __name__)


def _is_on_date(a: dict, date_str: str) -> bool:
    """
    created_at d'une analyse : datetime natif (MySQL local) OU chaîne ISO
    (backend Azure) → True si elle tombe sur date_str (AAAA-MM-JJ),
    sans TypeError ('datetime.datetime' object is not subscriptable).
    """
    c = a.get("created_at", "")
    day = c.strftime("%Y-%m-%d") if hasattr(c, "strftime") else str(c or "")[:10]
    return day == date_str


@cb_bp.get("/api/customerbalance/report")
@require_auth
def customerbalance_report():
    """
    Rapport spécifique pour le flux CustomerBalance.
    Lit n_integrated / n_rejected depuis le PairResult stocké en DB.
    """
    from datetime import datetime
    from storage import get_storage

    date_str = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))

    all_a = get_storage().list_analyses(limit=200)
    day_a = [
        a for a in all_a
        if _is_on_date(a, date_str)
        and any(kw in (a.get("flux_id", "") + a.get("flux_name", "")).upper()
                for kw in ["CUSTOMER", "CUSTOMERBALANCE", "CB"])
    ]

    pairs_summary = []
    for a in day_a:
        summary = a.get("summary", {})
        pairs = summary.get("pairs", [])
        for pair in pairs:
            pairs_summary.append({
                "label":             pair.get("label", "unknown"),
                "file_cegid":        pair.get("file_cegid", ""),
                "file_oracle":       pair.get("file_oracle", ""),
                "n_cegid_total":     int(pair.get("n_cegid", 0) or 0),
                "n_oracle_total":    int(pair.get("n_oracle", 0) or 0),
                "n_integrated":      int(pair.get("n_integrated", 0) or 0),
                "n_rejected":        int(pair.get("n_rejected", 0) or 0),
                "anomalies_count":   len(pair.get("anomalies", [])),
                "top_error_columns": pair.get("top_error_columns", []),
            })

    return jsonify({
        "date":       date_str,
        "flux_id":    "CUSTOMERBALANCE",
        "flux_name":  "Flux Customer Balance",
        "n_analyses": len(day_a),
        "pairs":      pairs_summary,
    })


@cb_bp.get("/api/report/customerbalance")
@require_auth
def report_customerbalance_json():
    """
    Endpoint JSON pour le rapport journalier CustomerBalance.
    Lit n_integrated / n_rejected directement depuis le PairResult en DB.
    Ces valeurs sont calculées dans generic_comparator.py (to_dict).
    """
    from datetime import datetime
    from storage import get_storage

    date_str = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))

    storage = get_storage()
    all_a = storage.list_analyses(limit=200)

    # Filtre par date ET par flux CustomerBalance
    day_a = [
        a for a in all_a
        if _is_on_date(a, date_str)
        and any(kw in (a.get("flux_id", "") + a.get("flux_name", "")).upper()
                for kw in ["CUSTOMER", "CUSTOMERBALANCE", "CB"])
    ]

    integrees      = 0
    rejetees       = 0
    n_total_cegid  = 0
    n_total_oracle = 0
    n_anomalies    = 0
    par_ou         = {}

    for a in day_a:
        summary = a.get("summary", {})
        pairs   = summary.get("pairs", [])
        for pair in pairs:
            # Ces valeurs sont maintenant présentes grâce au fix dans to_dict()
            ni  = int(pair.get("n_integrated", 0) or 0)
            nr  = int(pair.get("n_rejected",   0) or 0)
            nc  = int(pair.get("n_cegid",      0) or 0)
            no  = int(pair.get("n_oracle",     0) or 0)
            nan = len(pair.get("anomalies", []))
            ou  = pair.get("label") or pair.get("division") or "Global"

            integrees      += ni
            rejetees       += nr
            n_total_cegid  += nc
            n_total_oracle += no
            n_anomalies    += nan

            if ou not in par_ou:
                par_ou[ou] = {
                    "ou": ou, "integrees": 0, "rejetees": 0,
                    "n_cegid": 0, "n_oracle": 0, "n_anomalies": 0
                }
            par_ou[ou]["integrees"]   += ni
            par_ou[ou]["rejetees"]    += nr
            par_ou[ou]["n_cegid"]     += nc
            par_ou[ou]["n_oracle"]    += no
            par_ou[ou]["n_anomalies"] += nan

    total = integrees + rejetees

    return jsonify({
        "date":         date_str,
        "integrees":    integrees,
        "rejetees":     rejetees,
        "total":        n_total_cegid or total,
        "total_oracle": n_total_oracle,
        "n_anomalies":  n_anomalies,
        "taux_rejet":   round(rejetees / max(1, integrees + rejetees) * 100, 1),
        "par_division": list(par_ou.values()),
    })


@cb_bp.get("/api/report/customerbalance/csv")
@require_auth
def report_customerbalance_csv():
    """Export CSV du rapport CustomerBalance.
    
    Le rapport contient : Date, Division/OU, Lignes intégrées (CBLC1I), 
    Lignes rejetées (OPEC1R), Total Cegid, Taux rejet (%), Anomalies
    """
    from datetime import datetime
    import csv

    date_str = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))

    # --- Récupération directe des données (sans appel HTTP interne) ---
    try:
        storage = get_storage()
        all_a = storage.list_analyses(limit=200)
        from flask import current_app
        # Filtre les analyses du jour pour CustomerBalance
        analyses = [
            a for a in all_a
            if _is_on_date(a, date_str)
            and any(
                kw in (a.get("flux_id", "") + a.get("flux_name", "") + ".." + a.get("label", "")).upper()
                for kw in ["CUSTOMER", "CUSTOMERBALANCE", "CB"]
            )
        ]
    except Exception:
        analyses = []

    # Données par division/OU
    par_ou = {}
    integrees = 0
    rejetees = 0
    n_total = 0
    n_anomalies = 0
    for a in analyses:
        summary = a.get("summary", {})
        pairs = summary.get("pairs", [])
        for pair in pairs:
            ni = int(pair.get("n_integrated", 0) or 0)
            nr = int(pair.get("n_rejected", 0) or 0)
            nc = int(pair.get("n_cegid", 0) or 0)
            no = int(pair.get("n_oracle", 0) or 0)
            nan = len(pair.get("anomalies", []))
            ou = pair.get("label") or pair.get("division") or "Global"

            integrees += ni
            rejetees += nr
            n_total += nc or (ni + nr)
            n_anomalies += nan

            if ou not in par_ou:
                par_ou[ou] = {
                    "ou": ou, "integrees": 0, "rejetees": 0,
                    "n_cegid": 0, "n_anomalies": 0
                }
            par_ou[ou]["integrees"] += ni
            par_ou[ou]["rejetees"] += nr
            par_ou[ou]["n_cegid"] += nc
            par_ou[ou]["n_anomalies"] += nan

    # Construction du CSV
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "Date",
        "Division / OU",
        "Lignes intégrées (CBLC1I)",
        "Lignes rejetées (OPEC1R)",
        "Total Cegid",
        "Taux rejet (%)",
        "Anomalies"
    ])

    if par_ou:
        for d in par_ou.values():
            tot = d.get("integrees", 0) + d.get("rejetees", 0)
            tx = round(d.get("rejetees", 0) / max(1, tot) * 100, 1)
            writer.writerow([
                date_str,
                d.get("ou", "—"),
                d.get("integrees", 0),
                d.get("rejetees", 0),
                tot,
                tx,
                d.get("n_anomalies", 0)
            ])
    else:
        writer.writerow([date_str, "Toutes divisions", integrees, rejetees, n_total, "—", n_anomalies])

    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    response.headers["Content-Disposition"] = (
        f'attachment; filename="CustomerBalance_rapport_{date_str}.csv"'
    )
    return response


@cb_bp.get("/api/report/customerbalance/excel")
@require_auth
def report_customerbalance_excel():
    """Génère le rapport CustomerBalance en Excel (.xlsx) avec données réelles."""
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    date_str = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))

    # Récupère les données depuis l'endpoint JSON
    try:
        from flask import current_app
        with current_app.test_request_context(
            f"/api/report/customerbalance?date={date_str}",
            headers={"Authorization": request.headers.get("Authorization", "")}
        ):
            resp = report_customerbalance_json()
            data = resp.get_json()
    except Exception:
        data = {
            "integrees": 0, "rejetees": 0, "total": 0,
            "n_anomalies": 0, "taux_rejet": 0, "par_division": []
        }

    integrees    = data.get("integrees", 0)
    rejetees     = data.get("rejetees", 0)
    total_cegid  = data.get("total", integrees + rejetees)
    total_oracle = data.get("total_oracle", 0)
    n_anom       = data.get("n_anomalies", 0)
    divisions    = data.get("par_division", [])

    # ── Styles ──
    thin = Side(style="thin", color="CBD5E1")
    def brd():
        return Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    BLUE  = PatternFill("solid", start_color="1A46C8")
    RED   = PatternFill("solid", start_color="DC2626")
    GREEN = PatternFill("solid", start_color="16A34A")
    ORN   = PatternFill("solid", start_color="D97706")
    LGRE  = PatternFill("solid", start_color="DCFCE7")
    LRED  = PatternFill("solid", start_color="FEE2E2")
    LORN  = PatternFill("solid", start_color="FEF3C7")
    GREY  = PatternFill("solid", start_color="F1F5F9")

    def wf(bold=True, color="FFFFFF", size=11):
        return Font(name="Arial", bold=bold, color=color, size=size)

    wb = Workbook()

    # ════ Sheet 1 : Résumé ════
    ws = wb.active
    ws.title = "Résumé"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"Flux Monitor — CustomerBalance — {date_str}"
    c.font = wf(size=14); c.fill = BLUE; c.alignment = ctr; c.border = brd()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    c2 = ws["A2"]
    c2.value = "Direction : Oracle ➜ Cegid | Préfixe R (OPEC1R) = Rejeté, non intégré dans Cegid"
    c2.font = Font(name="Arial", bold=True, size=10, color="DC2626")
    c2.fill = LRED; c2.alignment = ctr; c2.border = brd()
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 10

    kpi_hdrs = [
        "Total lignes Cegid", "Total lignes Oracle",
        "Lignes intégrées (CBLC1I)", "Lignes rejetées (OPEC1R)",
        "Écart Cegid vs Oracle", "Nb anomalies total"
    ]
    kpi_vals = [total_cegid, total_oracle, integrees, rejetees,
                total_cegid - total_oracle, n_anom]
    kpi_fh = [BLUE, BLUE, GREEN, RED, ORN, RED]
    kpi_fv = [GREY, GREY, LGRE, LRED, LORN, LRED]

    for ci, (h, v, fh, fv) in enumerate(zip(kpi_hdrs, kpi_vals, kpi_fh, kpi_fv), 1):
        hc = ws.cell(4, ci, h)
        hc.font = wf(size=10); hc.fill = fh; hc.alignment = ctr; hc.border = brd()
        vc = ws.cell(5, ci, v)
        vc.font = Font(name="Arial", bold=True, size=16, color="1E293B")
        vc.fill = fv; vc.alignment = ctr; vc.border = brd()
        ws.column_dimensions[get_column_letter(ci)].width = 22

    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 42
    ws.row_dimensions[6].height = 10

    ws.merge_cells("A7:F7")
    st = ws["A7"]
    st.value = (
        f"STATUT : KO  |  Cegid={total_cegid} | Oracle={total_oracle} "
        f"| {rejetees} ligne(s) rejetée(s) | {n_anom} anomalies"
    )
    st.font = wf(size=11); st.fill = RED; st.alignment = ctr; st.border = brd()
    ws.row_dimensions[7].height = 28
    ws.row_dimensions[8].height = 14

    ws.merge_cells("A9:F9")
    lg = ws["A9"]
    lg.value = (
        "Légende : CBLC1I = Intégré dans Cegid et transmis à Oracle  |  "
        "OPEC1R = Rejeté (R), non intégré dans Cegid → invisible dans Oracle "
        "→ cause des anomalies MANQUANT_ORACLE"
    )
    lg.font = Font(name="Arial", size=10, color="1A46C8", bold=True)
    lg.fill = PatternFill("solid", start_color="EFF6FF")
    lg.alignment = lft; lg.border = brd()
    ws.row_dimensions[9].height = 32

    # ════ Sheet 2 : Rapport par OU ════
    ws2 = wb.create_sheet("Rapport par OU")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:F1")
    t2 = ws2["A1"]
    t2.value = f"Rapport par Division / OU — {date_str}"
    t2.font = wf(size=13); t2.fill = BLUE; t2.alignment = ctr; t2.border = brd()
    ws2.row_dimensions[1].height = 32

    cols_ou = [
        "Division / OU", "Lignes intégrées (CBLC1I)",
        "Lignes rejetées (OPEC1R)", "Total", "Taux rejet (%)", "Statut"
    ]
    for ci, h in enumerate(cols_ou, 1):
        c = ws2.cell(2, ci, h)
        c.font = wf(size=10); c.fill = BLUE; c.alignment = ctr; c.border = brd()
    ws2.row_dimensions[2].height = 28

    ou_data = divisions if divisions else []
    for ri, d in enumerate(ou_data, 3):
        int_v = d.get("integrees", 0)
        rej_v = d.get("rejetees", 0)
        tot   = int_v + rej_v
        tx    = round(rej_v / max(1, tot) * 100, 1)
        vals  = [
            d.get("ou", "—"), int_v, rej_v,
            f"=B{ri}+C{ri}", f"=IF(D{ri}=0,0,C{ri}/D{ri})",
            "⚠ KO" if rej_v > 0 else "✓ OK"
        ]
        fills = [None, LGRE, LRED, GREY, LORN, LRED if rej_v > 0 else LGRE]
        for ci, (val, fill) in enumerate(zip(vals, fills), 1):
            c = ws2.cell(ri, ci, val)
            c.border = brd(); c.alignment = ctr
            if fill: c.fill = fill
            c.font = Font(name="Arial", size=10, bold=(ci == 1))
            if ci == 5: c.number_format = "0.0%"
        ws2.row_dimensions[ri].height = 22

    tr = len(ou_data) + 3
    total_row = [
        "TOTAL",
        f"=SUM(B3:B{tr-1})", f"=SUM(C3:C{tr-1})", f"=SUM(D3:D{tr-1})",
        f"=IF(D{tr}=0,0,C{tr}/D{tr})", "⚠ KO"
    ]
    for ci, val in enumerate(total_row, 1):
        c = ws2.cell(tr, ci, val)
        c.font = wf(size=10); c.fill = BLUE; c.alignment = ctr; c.border = brd()
        if ci == 5: c.number_format = "0.0%"
    ws2.row_dimensions[tr].height = 26

    for ci, w in enumerate([30, 28, 28, 16, 18, 14], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    response = make_response(out.read())
    response.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response.headers["Content-Disposition"] = (
        f'attachment; filename="CustomerBalance_Rapport_{date_str}.xlsx"'
    )
    return response