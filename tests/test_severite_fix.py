"""Quick verification of the SEVERITE column fix for the ITEMS detail report."""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

import pandas as pd
from engine.detailed_report import build_detail_report, export_detailed_excel

data = {
    "ITEM_CODE":         ["A001", "A002", "A003", "A004"],
    "_merge":            ["both", "both", "left_only", "right_only"],
    "UNIT_PRICE_cegid":  ["10.00", "20.00", "30.00", None],
    "UNIT_PRICE_oracle": ["10.00", "25.00", None, "40.00"],
    "ITEM_BARCODE_cegid": ["BC1", "BC2", "BC3", None],
    "ITEM_BARCODE_oracle":["BC1", "BC2", None, "BC4"],
}
merged = pd.DataFrame(data)
cles = ["ITEM_CODE"]
valeurs = ["UNIT_PRICE", "ITEM_BARCODE"]

# 1. Without comparison_rules — ECART defaults to WARNING
report = build_detail_report(merged, cles, valeurs, "ITEMS")
print("=== Without comparison_rules ===")
for r in report:
    print(f"  {r['ITEM_CODE']}: STATUT={r['STATUT']}, SEVERITE='{r['SEVERITE']}'")
assert report[0]["SEVERITE"] == "", f"OK -> ''"
assert report[1]["SEVERITE"] == "WARNING", f"ECART default -> WARNING"
assert report[2]["SEVERITE"] == "WARNING", f"ABSENT_ORACLE -> WARNING"
assert report[3]["SEVERITE"] == "CRITIQUE", f"ABSENT_CEGID -> CRITIQUE"

# 2. With comparison_rules — ECART takes severity from rules
rules = [
    {"column": "UNIT_PRICE", "severity": "CRITIQUE", "tolerance": 0.01},
    {"column": "ITEM_BARCODE", "severity": "WARNING", "tolerance": 0},
]
report2 = build_detail_report(merged, cles, valeurs, "ITEMS", comparison_rules=rules)
print("\n=== With comparison_rules ===")
for r in report2:
    print(f"  {r['ITEM_CODE']}: STATUT={r['STATUT']}, SEVERITE='{r['SEVERITE']}'")
assert report2[0]["SEVERITE"] == ""
assert report2[1]["SEVERITE"] == "CRITIQUE", f"ECART on UNIT_PRICE -> CRITIQUE"
assert report2[2]["SEVERITE"] == "WARNING", f"ABSENT_ORACLE -> WARNING"
assert report2[3]["SEVERITE"] == "CRITIQUE", f"ABSENT_CEGID -> CRITIQUE"

# 3. Excel export: headers include SEVERITE after STATUT
import tempfile
path = export_detailed_excel(
    report2, {"nb_lignes_cegid": 4, "nb_lignes_oracle": 4}, "ITEMS",
    output_path=os.path.join(tempfile.gettempdir(), "test_severite.xlsx"),
)
from openpyxl import load_workbook
wb = load_workbook(path)

ws_detail = wb["Rapport détaillé"]
headers = [ws_detail.cell(row=1, column=c).value for c in range(1, ws_detail.max_column + 1)]
print(f"\n=== Excel headers: {headers} ===")
assert "SEVERITE" in headers
assert headers.index("SEVERITE") == headers.index("STATUT") + 1, "SEVERITE must be right after STATUT"

# 4. Écarts uniquement also has SEVERITE
ws_ecarts = wb["Écarts uniquement"]
ecarts_headers = [ws_ecarts.cell(row=1, column=c).value for c in range(1, ws_ecarts.max_column + 1)]
assert "SEVERITE" in ecarts_headers

# 5. Résumé tab counts from SEVERITE
ws_resume = wb["Résumé"]
critiques = warnings = None
for row in range(3, ws_resume.max_row + 1):
    label = ws_resume.cell(row=row, column=1).value
    val = ws_resume.cell(row=row, column=2).value
    if label == "Écarts critiques":
        critiques = int(val)
    elif label == "Écarts warnings":
        warnings = int(val)
    if label:
        print(f"  Résumé: {label} = {val}")

# ABSENT_CEGID(1) + ECART/CRITIQUE(1) = 2 critiques
# ABSENT_ORACLE(1) = 1 warning
assert critiques == 2, f"Expected 2 critiques, got {critiques}"
assert warnings == 1, f"Expected 1 warning, got {warnings}"

# 6. Verify coherence: counts from detail match Résumé
n_crit = sum(1 for r in report2 if r.get("SEVERITE") == "CRITIQUE")
n_warn = sum(1 for r in report2 if r.get("SEVERITE") == "WARNING")
assert critiques == n_crit, f"Résumé critiques ({critiques}) != detail ({n_crit})"
assert warnings == n_warn, f"Résumé warnings ({warnings}) != detail ({n_warn})"

print("\nALL TESTS PASSED")
